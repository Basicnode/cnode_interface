import json
from openpyxl import load_workbook

def parse_json_file(filepath):
    data = json.load(open(filepath, mode='r', encoding='utf8'))
    test_data = data["test_data"]
    return test_data

def parse_excel_file(filepath,sheetname):
    wb = load_workbook(filepath,sheetname)
    # print(wb.worksheets)
    ws = wb[sheetname]
    test_data = []
    for x in range(2, len(tuple(ws.rows)) + 1):
        testcase_data = []
        for y in range(2, 7):
            testcase_data.append(ws.cell(row=x, column=y).value)
            # print(ws.cell(row=x, column=y).value)
        test_data.append(testcase_data)
    return test_data


